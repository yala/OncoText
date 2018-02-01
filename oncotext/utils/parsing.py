from openpyxl import load_workbook
from collections import OrderedDict
import xmltodict
import datetime

def parse_XLS(path):
    wb = load_workbook(path)
    sheet = wb.active
    data = OrderedDict()
    for row in sheet.rows:
        values = [str(cell.value) for cell in row if cell.value is not None]
        if len(values) > 0:
            if len(values) > 1:
                data[values[0]] = values[1: ]
            else:
                data[values[0]] = []
    return data


def parse_XLS_reports(path, legend=None):
    wb = load_workbook(path)
    sheet = wb.active
    data = []
    legendRowBool = legend == None
    legend = [] if legend == None else legend
    for row in sheet.iter_rows():
        rowDict = {}
        for indx, col in enumerate(row):
            if legendRowBool:
                if not col.value in legend:
                    newName = col.value
                    legend.append(newName)
                else:
                    counter = 2
                    newName = str(counter) + " " + col.value
                    while newName in legend:
                        counter += 1
                        newName = str(counter) + " " + col.value
                    legend.append(newName)
            else:
                rowDict[legend[indx]] = col.value
        if not legendRowBool:
            data.append(rowDict)

        legendRowBool = False

        for d in reversed(data):
            if 'Report_Text' in d and d['Report_Text'] == None:
                data.remove(d)

            d['filename'] = path.split("/")[-1]

            for k in d:
                if isinstance(d[k], int):
                    try:
                        d[k] = str(d[k])
                    except Exception:
                        pass
                if isinstance(d[k], (datetime.datetime, datetime.date, datetime.time)):
                    d[k] = d[k].isoformat()

    return data


def parse_XML(path):
    xml = xmltodict.parse(open(path, 'rb'))
    dataroot = xml[list(xml.keys())[0]]
    reportsKey = [key for key in dataroot.keys() if ('@' not in key)][-1]
    annotatedReports = dataroot[reportsKey]
    annotatedReports = [r for r in annotatedReports if r!=None and len(r)>0]
    for r in annotatedReports:
        r['filename'] = path.split("/")[-1]

    return annotatedReports
